from __future__ import annotations

import hashlib
import json
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("openpyxl is required. Install it in the Python environment with: python -m pip install openpyxl")

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "Shopee_Keyword_Intelligence_Sprint2.xlsx"
OUTPUT_DIR = ROOT / "src" / "data" / "generated"
KEYWORDS_OUTPUT = OUTPUT_DIR / "keywords.json"
MANIFEST_OUTPUT = OUTPUT_DIR / "manifest.json"
CATEGORY_MAP_PATH = ROOT / "src" / "data" / "category-map.json"
SOURCE_SHEET = "Sprint 2 - Top 30"
SCHEMA_VERSION = 1
GENERATOR_VERSION = "1.0.0"
EXPECTED_HEADERS = ["Rank", "Keyword", "Cluster", "Category", "Intent", "Search Volume", "Score", "Priority", "Page Type", "SEO Title", "Product Research", "Target Products", "Selected Product 1", "Selected Product 2", "Selected Product 3", "Notes"]

def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", str(value))).strip()

def repair_source_text(value: Any) -> tuple[str, bool]:
    text = normalize_text(value)
    try:
        repaired = text.encode("cp1252").decode("cp874")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return text, False
    repaired = normalize_text(repaired)
    return repaired, repaired != text

def require_number(value: Any, field: str, row: int, *, integer: bool = False, minimum: float | None = None) -> int | float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"Row {row}: {field} must be numeric")
    if integer and int(value) != value:
        raise ValueError(f"Row {row}: {field} must be an integer")
    number: int | float = int(value) if integer else value
    if minimum is not None and number < minimum:
        raise ValueError(f"Row {row}: {field} must be at least {minimum}")
    return number

def make_slug(normalized_keyword: str, prefix: str) -> str:
    ascii_tokens = re.findall(r"[a-z0-9]+", normalized_keyword.casefold())
    readable = "-".join(ascii_tokens[:4])
    digest = hashlib.sha256(normalized_keyword.encode("utf-8")).hexdigest()[:10]
    return "-".join(part for part in (prefix, readable, digest) if part)

def canonical_hash(value: Any) -> str:
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(payload).hexdigest().upper()

def write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2) + "\n", encoding="utf-8", newline="\n")

def main() -> None:
    source_hash = hashlib.sha256(SOURCE.read_bytes()).hexdigest().upper()
    category_map = json.loads(CATEGORY_MAP_PATH.read_text(encoding="utf-8"))
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    try:
        if SOURCE_SHEET not in workbook.sheetnames:
            raise ValueError(f"Required sheet not found. Available sheets: {', '.join(workbook.sheetnames)}")
        sheet = workbook[SOURCE_SHEET]
        headers = [normalize_text(cell.value) if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        if headers != EXPECTED_HEADERS:
            raise ValueError(f"Unexpected headers in {SOURCE_SHEET}: {headers}")
        if workbook.properties.modified is None:
            raise ValueError("Workbook modified timestamp is required for deterministic generatedAt")
        generated_at = workbook.properties.modified.isoformat()
        records: list[dict[str, Any]] = []
        repair_count = 0
        for source_row, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(value not in (None, "") for value in values):
                continue
            raw = dict(zip(headers, values))
            repaired: dict[str, str] = {}
            for field in ("Keyword", "Cluster", "Category", "Intent", "Priority", "Page Type", "SEO Title", "Product Research"):
                if raw[field] in (None, ""):
                    raise ValueError(f"Row {source_row}: {field} is required")
                repaired[field], changed = repair_source_text(raw[field])
                repair_count += int(changed)
            rank = require_number(raw["Rank"], "Rank", source_row, integer=True, minimum=1)
            search_volume = require_number(raw["Search Volume"], "Search Volume", source_row, minimum=0)
            score = require_number(raw["Score"], "Score", source_row)
            category = repaired["Category"]
            if category not in category_map:
                raise ValueError(f"Row {source_row}: category has no explicit mapping status: {category}")
            normalized_keyword = normalize_text(repaired["Keyword"]).casefold()
            mapping = category_map[category]
            slug = make_slug(normalized_keyword, mapping["slugPrefix"])
            record = {
                "id": f"keyword-v1-{hashlib.sha256(normalized_keyword.encode('utf-8')).hexdigest()[:16]}",
                "rank": rank,
                "keyword": repaired["Keyword"],
                "normalizedKeyword": normalized_keyword,
                "slug": slug,
                "cluster": repaired["Cluster"],
                "category": category,
                "siteCategorySlug": mapping["siteCategorySlug"],
                "intent": repaired["Intent"],
                "pageType": repaired["Page Type"],
                "averageMonthlySearches": search_volume,
                "currency": None,
                "priorityScore": score,
                "priority": repaired["Priority"],
                "seoTitle": repaired["SEO Title"],
                "status": repaired["Product Research"],
                "sourceSheet": SOURCE_SHEET,
                "sourceRow": source_row,
                "generatedAt": generated_at,
                "schemaVersion": SCHEMA_VERSION,
                "sourceValues": {key: raw[key] for key in headers},
            }
            records.append(record)
    finally:
        workbook.close()
    records.sort(key=lambda record: record["rank"])
    unique_fields = ("rank", "keyword", "normalizedKeyword", "slug")
    for field in unique_fields:
        values = [record[field] for record in records]
        if len(values) != len(set(values)):
            raise ValueError(f"Duplicate {field} detected")
    unmapped = sorted({record["category"] for record in records if record["siteCategorySlug"] is None})
    warnings = []
    if repair_count:
        warnings.append(f"Repaired Windows-1252/CP874 mojibake in {repair_count} source text cells; raw values remain in sourceValues.")
    if "Currency" not in EXPECTED_HEADERS:
        warnings.append("Currency is not present in the source sheet; currency is null for every record.")
    if unmapped:
        warnings.append("Unmapped or review-required workbook categories: " + ", ".join(unmapped))
    keyword_hash = canonical_hash(records)
    manifest = {
        "schemaVersion": SCHEMA_VERSION,
        "sourceFile": "data/Shopee_Keyword_Intelligence_Sprint2.xlsx",
        "sourceFileSha256": source_hash,
        "sourceSheet": SOURCE_SHEET,
        "generatedAt": generated_at,
        "recordCount": len(records),
        "firstRank": records[0]["rank"] if records else None,
        "lastRank": records[-1]["rank"] if records else None,
        "keywordHash": keyword_hash,
        "contentHash": keyword_hash,
        "generatorVersion": GENERATOR_VERSION,
        "warnings": warnings,
    }
    write_json(KEYWORDS_OUTPUT, records)
    write_json(MANIFEST_OUTPUT, manifest)
    if hashlib.sha256(SOURCE.read_bytes()).hexdigest().upper() != source_hash:
        raise RuntimeError("Source workbook changed during import")
    print(f"Imported {len(records)} records from {SOURCE_SHEET}; keywordHash={keyword_hash}")

if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(f"Keyword import failed: {error}", file=sys.stderr)
        raise SystemExit(1)
